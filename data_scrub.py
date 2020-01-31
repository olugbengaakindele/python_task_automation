#this scripts scrubs excel files in different folders

import os
from os import walk
import openpyxl
import datetime


#lets set our all_data which will hold tthe data in a tabluar format but stored in pythin object first
avg_data = []
avg_record={}

st_time = datetime.datetime.now()
a=0
#set currebt workinf directory
path = (r"O:\Manufacturing\Production Control\Harvest Flower Cannabinoids QC Data")
os.chdir(path)

sub_fol = os.listdir(path)
#sub_fol returns a list of folders in the root folder eg roots_folder\2018,roots_folder\2019 etc
#we have to loop through the list of names
for x in range(0, len(sub_fol)):
    
    if os.path.isdir(sub_fol[x]):
        sub_fol_path = os.path.join(os.getcwd(),sub_fol[x])
        sub_fol_2 = os.listdir(sub_fol_path)
        
        #sub_fol_2 returns the list of ** folder and files and folders in roots/sub-folder/sub-folder-2
        for y in range(0, len(sub_fol_2)):
            sub_fol_2_path = os.path.join(os.getcwd(),sub_fol_path , sub_fol_2[y])
            #print(sub_fol_2_path)
            #if ** is a sub_folder then
            
            
            if os.path.isdir(sub_fol_2_path):
                s2_files = os.listdir(sub_fol_2_path)               
                #now , lets loop through the name os the files in the list 
                for s2 in range(0,len(s2_files)):
                    s2_file_path = os.path.join(os.getcwd(),sub_fol_path,sub_fol_2[y],s2_files[s2])
                    #s2_file_path gets the file path of a file so we can open it
                    #open the workbook
                    wb = openpyxl.load_workbook(s2_file_path)
                    all_sheets = wb.sheetnames
                    #all_sheets is a list of sheet names 
                    for all_s in all_sheets:
                        ws= wb[all_s]
                        loop_col = []
                        #lets loop through the columns
                        for col in range(1,ws.max_column):
                            if ws.cell(1,col).value:
                                col_value = ws.cell(1,col).value
                                if len(col_value) > 1 and ("pling" in col_value or "Canna" in col_value):
                                    loop_col.append(col)
                                         
                    #loop_gives us an arry of column to opp through to geth our variables
                    for xc in range(0,len(loop_col)):
                        ws.unmerge_cells(start_row=2,start_column=1,end_row=ws.max_row,end_column=ws.max_column)
                        'test to see if we are loopng through first column so that e get the averaget'
                        if xc == 0:
                            #loop through rows of first xolum in the loop col
                            for rows in range(1,ws.max_row):
                                #get the row that has the HLPC operator
                                if "HPLC Operator" in ws.cell(rows,loop_col[xc]):
                                    avg_record['HPLC Operator'] = "HPLC is in row " + rows
                                    avg_data.append(avg_record)
                                    print (avg_data)   
                        
            elif os.path.isfile(sub_fol_2_path):
                print(sub_fol_2_path)
                #open the workbook
                wb = openpyxl.load_workbook(sub_fol_2_path)
                all_sheets = wb.sheetnames
                #all_sheets is a list of sheet names 
                for all_s in all_sheets:
                    ws= wb[all_s] 
                    
                    loop_col = []
                    #lets loop through the columns
                    for col in range(1,ws.max_column):
                        if ws.cell(1,col).value:
                            col_value = ws.cell(1,col).value
                            if len(col_value) > 1 and ("pling" in col_value or "Canna" in col_value):
                                loop_col.append(col)
                    
                    print(loop_col)
                    #loop_gives us an arry of column to opp through to geth our variables
                    for xc in range(0,len(loop_col)):
                        ws.unmerge_cells(start_row=2,start_column=1,end_row=ws.max_row,end_column=ws.max_column)
                        'test to see if we are loopng through first column so that e get the averaget'
                        if xc == 0:
                            #loop through rows of first xolum in the loop col
                            for rows in range(1,ws.max_row):
                                #get the row that has the HLPC operator
                                if "HPLC Operator" in ws.cell(rows,loop_col[xc]):
                                    avg_record['HPLC Operator'] = "HPLC is in row " + rows
                                    avg_data.append(avg_record)
                                    print (avg_data)

end_time = datetime.datetime.now()   
print(end_time - st_time)

#conn = pyodbc.connect('Driver={SQL Server};Server=server_name;Database=db_nameo;Trusted_Connection=yes;')
#cursor = conn.cursor()
#cursor.execute('SELECT * FROM db_name.Table')

#for row in cursor:
 #   print(row)